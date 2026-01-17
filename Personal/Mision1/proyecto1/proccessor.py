import re
from openpyxl import load_workbook
# ====================================
# FUNCION Clean_id
# Elimina caracteres no numericos de un documento
# "cc75.888.56" = "7588856"
# ====================================
def clean_id(value):
    # Elimina caracteres no numericos de un documento
    if value is None:
        return ""
    return re.sub(r'\D', '',str(value))
# ====================================
#  FUNCION merge_name
#  Une nombre y apellido en un solo campo
# ====================================
def merge_name(name, lastname):
    if name is None:
        name =""
    if lastname is None:
        lastname =""
    return f"{name} {lastname}" .strip()

def proccess_excel(path):
    #Acceso a la hoja llamada "datos"
    wb = load_workbook(path)
    ws = wb["Datos"]
    #Recorrer todas las filas desde la fila 2
    for row in range(2,ws.max_row+1):
        #Columna D: identificador limpio
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        #Columna E: nombre completo
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value,
        )
    #Guarde los cambios en el mismo archivo
    wb.save(path)

def process_excel_safe(path):
    try:
        proccess_excel(path)
        return True, "Archivo procesado correctamente" 
    except PermissionError:
        return(
            False,
            "El archivo Excel está abierto.\n"
            "por Favor, Ciérrelo e intente nuevamente."
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False, f"Error inesperado: {str(e)}"
    
      
