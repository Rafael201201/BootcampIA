# processor.py
# Lógica de negocio: operaciones sobre Excel
import re
from openpyxl import load_workbook



def clean_id(value):
    # elimina carateres no numeros de un docummento
    if value is None:
        return""
    return re.sub(r"\D",'',str(value))
#========================================
#  funcion merge_name
#  une nombree y apellido en un solo campo
#====================================
def merge_name (name, lastname):
    if name is None:
        name=""
    if lastname is None:
        lastname=""
    return f"{name} {lastname}".strip()

def process_excel(path):
    # Acceso a la hoja de llamada"datos"
    wb= load_workbook(path)
    ws = wb["Datos"]
    # recorrer todas las filas desde la fila 2 
    for row in range (2,ws.max_row+1):
        #columnas D: identificador limpio
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo 
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value,
        )
        # guarde los cambios  en el mismo  archivo 
        wb.save(path)
        
def process_excel_safe(path):
    try:
        process_excel(path)
        print(path)
        return True,"archivo procesado correctamente"
    except PermissionError:
        return(
            False,
            "el archivo excel esta abierto.\n"
            "porfavor, cierrelo e intente nuevamente."
            )
    except KeyError:
        return False, "hoja'datos'no encontrada"
    except Exception as e:
        return False;f"error inesperado: {srt(e)}"
    
    
def ejecutar_accion(instruccion):
    # Abre el archivo de ejemplo
    wb = load_workbook("ejemplo.xlsx")
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(".xlsx")